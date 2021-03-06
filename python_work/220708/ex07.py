from openpyxl import load_workbook  # 파일열기는 load_workbook, 파일 생성은 workbook

wb = load_workbook("b.xlsx")
ws = wb.active

print('현재 작성된 행수 = ', ws.max_row)
print('현재 작성된 컬럼수 = ', ws.max_column)

for r in range(1,ws.max_row+1):
    for c in range(1,ws.max_column+1):
        print(f"r = {r}, c = {c} = ", ws.cell(row=r, column=c).value,end="")
    print()

wb.save("c.xlsx")
wb.close()



